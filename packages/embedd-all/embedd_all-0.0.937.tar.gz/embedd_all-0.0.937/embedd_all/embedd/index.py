import time
import openpyxl
import pandas as pd
from langchain.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
import uuid
from pinecone import Pinecone
from langchain.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter

from pinecone import ServerlessSpec
from tqdm import tqdm
import os
import json
import voyageai
from enum import Enum
from embedd_all.core.logger import logger
from docx import Document


class FileType(Enum):
    PDF = 'PDF'
    XLSX = 'XLSX'
    CSV = 'CSV'

def check_file_type(filepath):
    _, file_extension = os.path.splitext(filepath)
    if file_extension.lower() == '.pdf':
        return "PDF"
    elif file_extension.lower() == '.xlsx':
        return "XLSX"
    elif file_extension.lower() == '.csv':
        return "CSV"
    elif file_extension.lower() == '.doc' or file_extension.lower() == '.docx':
        return "DOC"
    else:
        return "Unknown format"
    
def check_metadata_size(metadata, limit=8000):
    """
    Check if metadata size exceeds a specified limit.
    
    :param metadata: Metadata to be checked.
    :param limit: Size limit in bytes.
    :return: Boolean indicating whether metadata size exceeds the limit.
    """
    metadata_size = len(json.dumps(metadata).encode('utf-8'))
    return metadata_size > limit



def find_column_row(df):
    final_index = None
    final_row = []
    for index, row in df.iterrows():
        # Convert all items to strings for consistent processing
        str_row = row.apply(lambda x: str(x) if pd.notna(x) else '')
        non_nan_count = str_row.str.strip().astype(bool).sum()
        unnamed_count = sum(col.startswith('Unnamed') for col in str_row if isinstance(col, str))

        if non_nan_count > unnamed_count and non_nan_count > 1:
            final_index = index
            final_row = row[row.notna()].tolist()
        
        if non_nan_count > unnamed_count and non_nan_count > 2:
            return index, row[row.notna()].tolist()
        
    return final_index, final_row

def unmerge_and_populate(file_path, sheet_name):
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Create a copy of the merged cell ranges
    merged_cells_copy = list(sheet.merged_cells.ranges)

    # Dictionary to hold cell values to populate
    cell_values = {}

    # Iterate through the copied merged cell ranges
    for merged_cell in merged_cells_copy:
        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
        
        # Check if the merged cells are vertical (same column)
        if min_col == max_col:
            top_cell_value = sheet.cell(row=min_row, column=min_col).value

            # Unmerge the cells
            sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

            # Store the value to populate in unmerged cells
            for row in range(min_row, max_row + 1):
                cell_values[(row, min_col)] = top_cell_value

    # Populate the unmerged cells with the stored values
    for (row, col), value in cell_values.items():
        sheet.cell(row=row, column=col, value=value)

    # Save the workbook
    workbook.save(file_path)

    # Load the sheet into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df

def convert_to_embedding_metadata(texts, embeddings):
    upsert_embeddings = []
    index = 0
    myuuid = uuid.uuid4()
    for text in texts:
        ob = {
            "id": str(myuuid),
            "values": embeddings[index],
            "metadata" : {
                "content": text
            }
        }
        index = index + 1
        upsert_embeddings.append(ob)
    return upsert_embeddings

# Define a function to create embeddings
def create_embeddings_voyage(texts, voyage_api_key, VOYAGE_EMBED_MODEL):
    vo = voyageai.Client(api_key=voyage_api_key)
    embeddings_list = []

    batch_size = 20
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i+batch_size]
        result = vo.embed(batch, model=VOYAGE_EMBED_MODEL, input_type="document")
        embeddings_list.extend(result.embeddings)

    # embeds = sum(embeddings_list, [])
    return embeddings_list


def pinecone_embeddings_with_voyage_ai(paths, pinecone_key, voyage_api_key, vector_db_name, voyage_embed_model, embed_dimension):
    pc = Pinecone(api_key=pinecone_key)

    cloud = 'aws'
    region = 'us-east-1'

    spec = ServerlessSpec(cloud=cloud, region=region)
    VOYAGE_EMBED_MODEL = voyage_embed_model

    index_name = vector_db_name
    # check if index already exists (it shouldn't if this is first time)
    if index_name not in pc.list_indexes().names():
        # if does not exist, create index
        pc.create_index(
                index_name,
                dimension=embed_dimension,
                metric='cosine',
                spec=spec
        )
        # wait for index to be initialized
        while not pc.describe_index(index_name).status['ready']:
            time.sleep(1)

        # connect to index
    pc_index = pc.Index(index_name)
        # view index stats
    pc_index.describe_index_stats()

    base_path = ''

    for path in paths:
        file_type = check_file_type(path)
        file_path = base_path + path
        texts = []

        if file_type == "PDF":
            logger.info(f"PDF processing started ... {path}")
            texts = process_pdf(file_path)
            logger.info(f"PDF processing complete ... {path}")

        if file_type == "XLSX":
            context = "data"
            logger.info(f"EXCEL processing started ... {path}")
            dfs = modify_excel_for_embedding(file_path=file_path, context=context)
            texts = [text for df in dfs for text in df]
            logger.info(f"EXCEL processing complete ... {path}")
        
        if file_type == "CSV":
            logger.info(f"CSV processing started ... {path}")
            context = "data"
            dfs = modify_csv_for_embedding(file_path=file_path, context=context)
            texts = [text for df in dfs for text in df]
            logger.info("fCSV processing complete ... {path}")

        if file_type == "DOC":
            logger.info(f"DOC processing started ... {path}")
            texts = read_docx_as_pages(file_path)
            logger.info(f"DOC processing complete ... {path}")

        if not texts:
            logger.info(f"No texts extracted from {path}. Skipping embedding process.")
            continue

        embeddings = create_embeddings_voyage(texts, voyage_api_key, VOYAGE_EMBED_MODEL)
        logger.info(f"Embeddings Length: {len(embeddings)}")
        logger.info(f"Embedding complete ... {path}")
        upsert_embeds = convert_to_embedding_metadata(texts, embeddings)
        logger.info(f"Upsert Embed Length: {len(upsert_embeds)}")
        logger.info(f"Embedding to pine cone started ... {path}")
        batch_size = 5
        # Wrap the range in tqdm for progress tracking
        for i in tqdm(range(0, len(upsert_embeds), batch_size), desc="Upserting batches", unit="batch"):
            batch = upsert_embeds[i:i+batch_size]
            pc_index.upsert(vectors=batch)
        
        logger.info(f"Embedding to pine cone complete ... {path}")

    logger.info(f"Embedding Completed")

def process_pdf(file_path):
    # create a loader
    loader = PyPDFLoader(file_path)
    # load your data
    data = loader.load()
    # Split your data up into smaller documents with Chunks
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=0)
    documents = text_splitter.split_documents(data)
    # Convert Document objects into strings
    texts = [str(doc) for doc in documents]
    return texts

def modify_excel_for_embedding(file_path, context):
    dfs = []

    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names
    logger.info(f"Sheet names: {sheet_names}")
    
    for sheet_index, sheet in enumerate(sheet_names, 1):
        df = unmerge_and_populate(file_path, sheet)

        # Find the row with column names
        column_row_index, column_names = find_column_row(df)
        if column_row_index is not None:
            logger.info(f"Column names: {column_names}")
        else:
            logger.info("No valid column row found.")

        # If column names are found, set the column names and drop the rows above it
        if column_row_index is not None:
            # Fill the beginning NaNs with placeholder column names
            column_names = ['Unnamed: ' + str(i) if pd.isna(col) else col for i, col in enumerate(df.iloc[column_row_index])]
            df.columns = column_names
            df = df.drop(range(column_row_index + 1)).reset_index(drop=True)
            
            # Drop columns starting with "Unnamed: "
            df = df.loc[:, ~df.columns.astype(str).str.contains('Unnamed: ')]

        columns = df.columns.tolist()

        # Initialize the "summarized" column
        df["summarized"] = ""

        # Iterate through each row to summarize the data
        for index, row in df.iterrows():
            summary = []
            
            for col in columns:
                value = row[col]
                if pd.notna(value):  # Only include non-null values
                    summary.append(f"{col}: {str(value).strip()}")
            
            summary_str = "; ".join(summary)
            if sheet_names:
                summary_str = f"{sheet}/{summary_str}"
            df.at[index, "summarized"] = f"{context}/{summary_str}"

        dfs.append(df["summarized"])
        logger.info(f"Processed Sheet: {sheet_index}")

    return dfs

def modify_csv_for_embedding(file_path, context):
    # Read the CSV file
    df = pd.read_csv(file_path, header=None)
    # Find the row with column names
    column_row_index, column_names = find_column_row(df)
    if column_row_index is not None:
        logger.info(f"Column names: {column_names}")
    else:
        logger.info("No valid column row found.")
        return []

    # Set the column names and drop the rows above it
    df.columns = column_names
    df = df.iloc[column_row_index + 1:].reset_index(drop=True)
    
    # Drop columns starting with "Unnamed: "
    df = df.loc[:, ~df.columns.astype(str).str.contains('Unnamed: ')]

    columns = df.columns.tolist()

    # Initialize the "summarized" column
    df["summarized"] = ""

    # Iterate through each row to summarize the data
    for index, row in df.iterrows():
        summary = []
        
        for col in columns:
            value = row[col]
            if pd.notna(value):  # Only include non-null values
                summary.append(f"{col}: {str(value).strip()}")
        
        summary_str = "; ".join(summary)
        df.at[index, "summarized"] = f"{context}/{summary_str}"

    logger.info(f"Processed CSV file")
    return [df["summarized"]]


def read_docx_as_pages(file_path, size_limit=8000):
    # Open the .docx file
    doc = Document(file_path)

    # Initialize an empty list to hold the text "pages"
    pages = []
    
    # Iterate over each paragraph and group them into "pages"
    current_page = []
    for para in doc.paragraphs:
        current_page.append(para.text)

        if check_metadata_size("\n".join(current_page), size_limit):
            pages.append("\n".join(current_page[:-1]))
            current_page = [current_page[-1]]
        
        # Treat every 23 paragraphs as a new "page" (adjust as needed)
        # if len(current_page) >= 1:
        #     pages.append("\n".join(current_page))
        #     current_page = []

    # Append any remaining paragraphs as the last "page"
    if current_page:
        pages.append("\n".join(current_page))

    return pages