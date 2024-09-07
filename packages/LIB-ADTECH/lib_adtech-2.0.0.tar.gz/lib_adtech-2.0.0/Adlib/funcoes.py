import requests
import datetime

def importarLibs():
    import pandas as pd
    from typing import Text
    from bs4 import element
    from bs4.element import ProcessingInstruction
    from numpy import SHIFT_DIVIDEBYZERO, False_, exp
    from pandas.io import html
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    import time
    import pandas as pd
    from pandas.core.frame import DataFrame
    from pandas.io.parsers import read_csv
    from selenium import webdriver
    from selenium.webdriver.chrome.webdriver import WebDriver
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.support import expected_conditions
    from selenium.webdriver.support.wait import WebDriverWait
    from selenium.webdriver.chrome.options import Options
    from bs4 import BeautifulSoup, BeautifulStoneSoup
    from selenium import webdriver
    from openpyxl import Workbook, load_workbook
    import os
    import tkinter as tk
    from tkinter import messagebox
    from discord_webhook import DiscordWebhook, DiscordEmbed
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
    # Importe o datetime aqui
    from datetime import datetime

    bibliotecas = {
        'os.path': os.path,
        'pd': pd,
        'Text': Text,
        'element': element,
        'ProcessingInstruction': ProcessingInstruction,
        'SHIFT_DIVIDEBYZERO': SHIFT_DIVIDEBYZERO,
        'False_': False_,
        'exp': exp,
        'html': html,
        'webdriver': webdriver,
        'Keys': Keys,
        'ActionChains': ActionChains,
        'time': time,
        'DataFrame': DataFrame,
        'read_csv': read_csv,
        'Options': Options,
        'BeautifulSoup': BeautifulSoup,
        'BeautifulStoneSoup': BeautifulStoneSoup,
        'Workbook': Workbook,
        'load_workbook': load_workbook,
        'datetime': datetime,
        'tk': tk,
        'messagebox': messagebox,
        'DiscordWebhook': DiscordWebhook,
        'DiscordEmbed': DiscordEmbed,
        'WebDriverWait': WebDriverWait,
    }

    return bibliotecas

def esperarElemento(driver, valor, tempo_espera=10):
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.keys import Keys
        return WebDriverWait(driver, tempo_espera).until(EC.visibility_of_element_located(('xpath', valor)))

def clickarElemento(driver, valor, time_wait=10):
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    return WebDriverWait(driver, time_wait).until(EC.element_to_be_clickable(('xpath', valor)))

def finalizarPorHora(hora, minutes):
    agora = datetime.now()
    t = agora.strftime('%H:%M')
    t2 = agora.replace(hour=hora, minute=minutes)
    t_simplificado = t2.strftime('%H:%M')

    return t >= t_simplificado

def mensagemTelegam(token, chat_id, mensagem):
    mensagem_formatada = f'https://api.telegram.org/bot{token}/sendMessage?chat_id={chat_id}&parse_mode=HTML&text={mensagem}'
    resposta = requests.get(mensagem_formatada)
    return resposta.json()






  


