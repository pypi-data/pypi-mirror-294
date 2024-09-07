from copy import copy

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    DEFAULT_FONT,
    NamedStyle,
    Alignment,
)
from openpyxl.styles.borders import DEFAULT_BORDER
from openpyxl.styles.fills import FILL_SOLID

# OVERRIDE default font
DEFAULT_FONT.size = 10
DEFAULT_FONT.name = 'Calibri'

align_left = Alignment(horizontal='left')

fat_top_border = copy(DEFAULT_BORDER)
fat_top_border.top = Side(border_style='thick', color='575757')
thin_top_border = copy(DEFAULT_BORDER)
thin_top_border.top = Side(border_style='thin', color='575757')

h1 = NamedStyle(name='h1')
h1.fill = PatternFill(fill_type=FILL_SOLID, start_color='d9d9d9', end_color='d9d9d9')
h1.font = copy(DEFAULT_FONT)
h1.font.color = '575757'
h1.font.size = 14
h1.font.bold = True
h1.border = copy(fat_top_border)

h1sub = NamedStyle(name='h1-sub')
h1sub.fill = copy(h1.fill)
h1sub.font = copy(h1.font)
h1sub.font.size = 11
h1sub.font.bold = False

h2 = NamedStyle(name='h2')
h2.fill = copy(h1.fill)
h2.font = copy(h1.font)
h2.font.size = 12
h2.border = thin_top_border

h2sub = NamedStyle(name='h2-sub')
h2sub.fill = copy(h2.fill)
h2sub.font = copy(h2.font)
h2sub.font.bold = True

h3 = NamedStyle(name='h3')
h3.fill = copy(h2.fill)
h3.font = copy(h2.font)
h3.font.size = 10
h3.border = thin_top_border

h3sub = NamedStyle(name='h3-sub')
h3sub.fill = copy(h3.fill)
h3sub.font = copy(h3.font)
h3sub.font.bold = False


TABLE_HEADER_FILL = PatternFill(
    fill_type=FILL_SOLID, start_color='575757', end_color='575757'
)
TABLE_HEADER_FONT = Font('Calibri', size=10, bold=True, color='FFFFFF')
TABLE_FILL_ALTERNATE = PatternFill(
    fill_type=FILL_SOLID, start_color='EEEEEE', end_color='EEEEEE'
)
HEADER_FONT = Font('Calibri', size=10, bold=True)
thin = Side(border_style='thin', color='b0b0b0')
BORDER = Border(top=thin, left=thin, right=thin, bottom=thin)
black = Side(border_style='thin', color='000000')
BORDER_BLACK = Border(top=black, left=black, right=black, bottom=black)

FMT_GENERAL = 'General'
FMT_EUR = '#,##0.00 €'
FMT_HOUR = '#,##0.00" h"'
FMT_KG = '#,##0.00" kg"'
FMTS = {
    'Arbeit': FMT_EUR,
    'Aufwand': FMT_HOUR,
    'Dauer': FMT_HOUR,
    'Einzelpreis': FMT_EUR,
    'Ges.gew.': FMT_KG,
    'Ges.preis': FMT_EUR,
    'Gesamt': FMT_EUR,
    'Gesamtgewicht': FMT_KG,
    'Gesamtpreis': FMT_EUR,
    'Gewicht': FMT_KG,
    'Material': FMT_EUR,
    'Materialpreis': FMT_EUR,
    'fixer Aufwand': FMT_HOUR,
    'var. Aufwand': FMT_HOUR,
}
