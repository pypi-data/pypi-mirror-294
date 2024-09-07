import string
from pathlib import Path

import openpyxl.utils
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet

from . import fmt
from .edi import EDISTR, EDI
from .table import Table


class Sheet:
    def __init__(self, ws: Worksheet, fmts: dict = None):
        self.ws = ws
        self.fmts = fmts
        if self.fmts is None:
            self.fmts = fmt.FMTS

    def add_table(self, start: EDISTR, table: Table, fmts: dict = None) -> EDI:
        if fmts is None:
            fmts = {}
        start = EDI(start)
        my_formats = self.fmts | fmts
        formats = [my_formats.get(c) for c in table.header]
        for edi, v in start.with_col_indices(table.header).items():
            self.ws[edi] = v
            self.ws[edi].font = fmt.TABLE_HEADER_FONT
            self.ws[edi].fill = fmt.TABLE_HEADER_FILL
            self.ws[edi].border = fmt.BORDER
        for row_i, row in enumerate(table.data):
            stripe_row = row_i % 2 == 0
            data_start = start.row_offset(row_i + 1)  # +1 to jump over header
            for col_i, (edi, v) in enumerate(data_start.with_col_indices(row).items()):
                self.ws[edi] = v
                self.ws[edi].border = fmt.BORDER
                if stripe_row:
                    self.ws[edi].fill = fmt.TABLE_FILL_ALTERNATE
                x = formats[col_i]
                if x is not None:
                    self.ws[edi].number_format = x
        return start.row_offset(len(table.data) + 1)

    def add_table_top_down(
        self, start: EDISTR, table: Table, merge_cols: int = 2
    ) -> EDI:
        """
        :merge_cols: Number of columns to merge for values.
        :return: Row index after inserted data, e.g. B11.
        """
        start = EDI(start)
        d = table.single_dict()
        # iterate rows
        for i, (key, value) in enumerate(d.items()):
            header_edi = start.row_offset(i)
            value_edi = start.row_offset(i).col_offset(1)
            self.ws[header_edi] = key
            self.ws[header_edi].font = fmt.TABLE_HEADER_FONT
            self.ws[header_edi].fill = fmt.TABLE_HEADER_FILL
            self.ws[header_edi].border = fmt.BORDER
            self.ws[value_edi] = value
            self.ws[value_edi].border = fmt.BORDER
            self.ws[value_edi].number_format = self.fmts.get(key, fmt.FMT_GENERAL)
            self.ws[value_edi].alignment = fmt.align_left
            # merge cells
            self.ws.merge_cells(value_edi.get_range_col_offset(merge_cols))
        return start.row_offset(len(d))

    def add_table_with_header(
        self,
        start: EDISTR,
        header: str,
        table: Table,
        style: str,
        end_col='J',
        fmts: dict = None,
    ):
        start = EDI(start)
        self.ws[start] = header
        self.style_row(start, style, end_col)
        return self.add_table(start.row_offset(1), table, fmts=fmts)

    # noinspection DuplicatedCode
    def style_row(self, start: EDISTR, style: str, end_col: str = 'J'):
        start = EDI(start)
        a, b = (
            string.ascii_uppercase.index(start.col),
            string.ascii_uppercase.index(end_col) + 1,
        )
        cols = string.ascii_uppercase[a:b]
        for c in cols:
            self.ws[start.at_col(c)].style = style
        return start.row_offset(1)

    # noinspection DuplicatedCode
    def fill_row(self, start: EDISTR, end_col: str, fill):
        start = EDI(start)
        a, b = (
            string.ascii_uppercase.index(start.col),
            string.ascii_uppercase.index(end_col),
        )
        cols = string.ascii_uppercase[a:b]
        for c in cols:
            self.ws[start.at_col(c)].fill = fill

    def auto_set_column_width(self, row_start=1, min_width=10):
        """
        https://stackoverflow.com/a/58667753/241240
        """
        for i, column_cells in enumerate(self.ws.columns, start=1):
            # note that black formats this weirdly. Just live with it. :P
            width = (
                length
                if (
                    length := max(
                        len(
                            str(cell_value)
                            if (cell_value := cell.value) is not None
                            else ''
                        )
                        for cell in column_cells[(row_start - 1) :]  # noqa
                    )
                )
                >= min_width
                else min_width
            )
            # + 1 wegen units (kg)
            self.ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = (
                width + 4
            )

    def add_logo(self, path: Path):
        img = Image(path)
        img.anchor = 'B1'
        img.height = 79
        img.width = 162
        self.ws.add_image(img)

    def width_in_inch(self, column, inches):
        x = inches * 15.8
        self.ws.column_dimensions[column].width = x

    def set_widths_in_inch(self, col2inch: dict[str, float]):
        for col, inch in col2inch.items():
            self.width_in_inch(col, inch)
