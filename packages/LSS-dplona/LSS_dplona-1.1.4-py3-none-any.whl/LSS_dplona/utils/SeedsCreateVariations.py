# Date:     01 May 2018
# File:     CreateVariations.py
# By:       JBB
# Subject:  Python script to create variation OrcaFlex files based on 
#           input from Excelsheet
# 
# Input:    Values in an Excelsheet
#           The references to the Excelsheet and the base case model need to 
#           be adjusted in this file
#           Also the parameters need to be changed case-by-case and the what 
#           to do with the parameters in OrcaFlex
# 
# Output:   OrcaFlex files with the variations

import os.path
import shutil
import openpyxl
import OrcFxAPI

# os.path — Common pathname manipulations (e.g. check if a directory exists)
# itertools — Functions creating iterators for efficient looping (e.g. 
#             create a list with all possible variations of input parameters)
# shutil — High-level file operations (e.g. remove entire directory)
# yaml - Used to write comments to the OrcaFlex file
# openpyxl - A Python library to read/write Excel 2010 xlsx/xlsm files
# OrcFxAPI - The OrcaFlex API
# math — Mathematical functions

# Definition of constants for this file

def ReadColumn(sheet, column, title):
# This function reads columns in the worksheet
# It checks if the name in the header row is equal to the expected 
# parameter name
# It reads values in a column until it encounters an empty cell
    header = sheet.cell(row=1, column=column).value
    # != is the Python is not operator
    if title != header:
	    raise ValueError('Unexpected column header, expected {0}, was {1}'
                      .format(title, header))
    values = []
    row = 2
    while True:
        value = sheet.cell(row=row, column=column).value
        if value is None:
            break
        values.append(value)
        row += 1
    print(title, '=', values)
    return values

def ReadNumericColumn(sheet, column, title):
# This function converts the output from ReadColumn to a floating point number
    return [float(value) for value in ReadColumn(sheet, column, title)]

def CreateVariationFiles(baseModel, outputDir, H, T, D, nSeed, minT):
# This function creates the models with all possible variations of 
# input parameters
# It also writes a filename.txt to keep track of all files created 
    varModel = OrcFxAPI.Model(threadCount=1)
    varModel.LoadData(baseModel)
    
    caseNum = 1
    
    fileNames = []
    fileCount = len(H)*50
    caseNumFmt = '0{0}d'.format(len(str(fileCount)))
    fileNameFmt = '#{0:' + caseNumFmt + '},Hs={1},Tp={2},Dir={3},Seed={4}.dat'

    env = varModel.environment    
    fmax = env.WaveSpectrumMaxRelFrequency
    
    for i in range(0,len(H)):
        for j in range (nSeed):
            h = H[i]
            t = T[i]
            d = D[i]

            env.SelectedWave = 'Wave1'
            env.WaveHs = h
            env.WaveTp = t
            env.WaveDirection = d
            env.WaveSeed = j
            
            
            
            if 1/t*fmax > 1/minT:
                env.WaveSpectrumMaxRelFrequency = (1/minT) / (1/t)
            
            # save the file
            fileName = fileNameFmt.format(caseNum, h, float(t), float(d), j)
            fileNames.append(fileName)
            varModel.SaveData(os.path.join(outputDir, fileName))

            print('Saved:', fileName)
            caseNum += 1
            
if __name__ == '__main__':    
        
    outputDir = 'Iteration_0_new'
    
    if os.path.isdir(outputDir):
        shutil.rmtree(outputDir)
    os.mkdir(outputDir)
    
    # data_only=True ensures that we evaluate rather than read formulae
    book = openpyxl.load_workbook('InputVariations.xlsx', 
                                  data_only=True)
    
    # Read the variation parameters
    print('Read variation parameters:')
    sheet = book['Parameters']
    H = ReadNumericColumn(sheet, 1, 'hs')
    T = ReadNumericColumn(sheet, 2, 'tp')
    D = ReadNumericColumn(sheet, 3, 'dir')
    print('\n')
    
    # Use of threadCount=1 avoids creating threads that are never used
    model = 'Splashzone lowering.dat'
    nSeed = 40
    minT = 1
    CreateVariationFiles(model, outputDir, H, T, D, nSeed, minT)
    


